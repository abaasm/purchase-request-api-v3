from flask import Flask, request, send_file, jsonify
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
import tempfile

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

def generate_purchase_request(df, supplier_name, months_of_cover, months_to_average=None):
    """
    Generate a purchase request for a specific supplier with product image formulas.
    """
    # Normalize column names
    df_work = df.copy()
    df_work.columns = [str(c).strip().lower() for c in df_work.columns]
    
    # Identify month columns (exclude known non-month columns)
    non_month_cols = set(['products', 'current stock', 'supplier', 'lc'])
    all_month_cols = [c for c in df_work.columns if c not in non_month_cols]
    
    # If specific months specified, use those; otherwise use all month columns
    if months_to_average is not None:
        months_to_use = [str(m).strip().lower() for m in months_to_average]
    else:
        months_to_use = all_month_cols
    
    # Ensure numeric for month columns and current stock
    for c in months_to_use:
        if c in df_work.columns:
            df_work[c] = pd.to_numeric(df_work[c], errors='coerce')
    df_work['current stock'] = pd.to_numeric(df_work['current stock'], errors='coerce')
    
    # Handle LC column if present
    if 'lc' in df_work.columns:
        df_work['lc'] = pd.to_numeric(df_work['lc'], errors='coerce')
    
    # Compute average monthly sales over specified months
    df_work['avg_monthly_sales'] = df_work[months_to_use].mean(axis=1, skipna=True)
    
    # Filter to supplier
    req_df = df_work[df_work['supplier'].astype(str).str.strip() == supplier_name.strip()].copy()
    
    if len(req_df) == 0:
        return None, months_to_use
    
    # Fill NaN values with 0 for calculations
    req_df['avg_monthly_sales'] = req_df['avg_monthly_sales'].fillna(0)
    req_df['current stock'] = req_df['current stock'].fillna(0)
    
    # Compute required quantity
    req_df['target_stock'] = req_df['avg_monthly_sales'] * months_of_cover
    req_df['shortfall'] = req_df['target_stock'] - req_df['current stock']
    
    # Handle NaN and inf before converting to int
    req_df['shortfall'] = req_df['shortfall'].replace([np.inf, -np.inf], 0).fillna(0)
    req_df['purchase_qty'] = np.where(req_df['shortfall'] > 0, np.ceil(req_df['shortfall']), 0).astype(int)
    
    # Add image formulas
    req_df['product_image'] = req_df['products'].apply(
        lambda x: '=IMAGE("https://ecomedia.shaghlaty.net/media/catalog/' + str(x) + '.jpg")'
    )
    
    # Select output columns - include LC columns if present
    out_cols = ['products', 'product_image', 'supplier'] + months_to_use + ['avg_monthly_sales', 'current stock', 'target_stock', 'purchase_qty']
    
    # Add LC columns if LC exists in the data
    if 'lc' in req_df.columns:
        req_df['lc_in_usd'] = req_df['lc'].fillna(0)
        req_df['lc_in_iqd'] = req_df['lc'].fillna(0) * 1550
        out_cols.extend(['lc_in_usd', 'lc_in_iqd'])
    
    output = req_df[out_cols].sort_values('purchase_qty', ascending=False)
    
    return output, months_to_use

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        'service': 'Shaghlaty Purchase Request API',
        'version': '1.1',
        'endpoints': {
            '/generate_purchase_request': {
                'method': 'POST',
                'description': 'Generate a purchase request Excel file',
                'parameters': {
                    'file': 'Excel file (multipart/form-data)',
                    'supplier_name': 'Name of supplier (form field)',
                    'months_of_cover': 'Number of months to cover (form field, integer)',
                    'months_to_average': 'Optional: comma-separated month columns to average (form field)'
                },
                'returns': 'Excel file download with LC in USD and LC in IQD columns (if LC column present in input)'
            }
        }
    })

@app.route('/generate_purchase_request', methods=['POST'])
def generate_purchase_request_endpoint():
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Get parameters
        supplier_name = request.form.get('supplier_name')
        months_of_cover = request.form.get('months_of_cover')
        months_to_average = request.form.get('months_to_average', None)
        
        if not supplier_name:
            return jsonify({'error': 'supplier_name is required'}), 400
        if not months_of_cover:
            return jsonify({'error': 'months_of_cover is required'}), 400
        
        try:
            months_of_cover = int(months_of_cover)
        except ValueError:
            return jsonify({'error': 'months_of_cover must be an integer'}), 400
        
        # Parse months_to_average if provided
        if months_to_average:
            months_to_average = [m.strip() for m in months_to_average.split(',')]
        else:
            months_to_average = None
        
        # Read the Excel file
        df = pd.read_excel(file)
        
        # Generate purchase request
        output, months_used = generate_purchase_request(df, supplier_name, months_of_cover, months_to_average)
        
        if output is None:
            return jsonify({'error': 'Supplier not found in the inventory file'}), 404
        
        # Create temporary file for output
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        output_path = temp_file.name
        temp_file.close()
        
        # Save to Excel
        output.to_excel(output_path, index=False)
        
        # Adjust row heights and column width using openpyxl
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Set row height to 100 for all rows
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 100
        
        # Widen the image column (column B)
        ws.column_dimensions['B'].width = 25
        
        wb.save(output_path)
        
        # Generate filename
        safe_supplier = secure_filename(supplier_name)
        output_filename = 'purchase_request_' + safe_supplier + '_' + str(months_of_cover) + 'months.xlsx'
        
        # Send file
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
